import io
import os
import re
import tempfile
from collections import Set, deque

import datetime

import pickle
from numpy import random

import xlsxwriter
from flask import flash
from flask import request
from flask import send_file
from flask.templating import render_template_string
from openpyxl import load_workbook
from sqlalchemy import cast
from sqlalchemy import inspect
from werkzeug.utils import secure_filename
from wtforms import Form, validators, FileField
from wtforms.fields.core import IntegerField, SelectField, StringField, BooleanField
from wtforms.fields.simple import TextAreaField, HiddenField

import ehbmail
from __init__ import *
from config import read_songs
from extras import extras_cost
from tables import *
from helpers import *
from flask_login import login_required, current_user
from discount import *
from add_geolocations import *
from generate_map import *

delay_between_messages = float(conf["email"]["delay_between_messages"])


@app.route('/admin.html')
@login_required
def adminpage():
    return render_template("admin.html")


@app.route('/show-log.html')
@login_required
def show_logfile():
    # read log file, possibly cropping it to the last N lines
    with open(log_file_name, 'r', encoding="utf-8") as myfile:
        if conf.has_option("server", "logfile_max_lines"):
            all_lines = deque(myfile, maxlen=conf.getint("server", "logfile_max_lines"))
        else:
            all_lines = list(myfile)

    # parse log entries
    items = []
    for line in all_lines:
        # 2017-05-01 09:22:25,473 INFO	Tornado started at 2017-05-01 09:22:23
        mat = re.match('^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d{3})\s+(\S+)\s+(.*)$', line)
        if mat is not None:
            timestamp = mat.group(1)
            level = mat.group(2)
            message = mat.group(3)
        else:
            timestamp = ""
            level = "INFO"
            message = line

        items.append((timestamp, level, message))

    return render_template("logfile.html", items=items)


@app.route("/mailtool.html", methods=["GET", ])
@login_required
def mailtool():
    form = MailtoolForm(request.form)
    prts = session.query(Participant).all()
    return render_template("mailtool.html", title="Mail Tool", form=form, participants=prts, delay=delay_between_messages)


@app.route("/mailtool.html", methods=["POST", ])
@login_required
def do_mailtool():
    form = MailtoolForm(request.form)

    if form.validate():
        submit_button_label = request.form['submit']

        if not form.recipients.data:
            # no recipients selected
            prts = session.query(Participant).all()
            return render_template("mailtool.html", title="Mail Tool", form=form, participants=prts, delay=delay_between_messages)

        if submit_button_label == "revise":
            # "revise" button clicked on confirm page
            prts = session.query(Participant).all()
            return render_template("mailtool.html", title="Mail Tool", form=form, participants=prts, delay=delay_between_messages)

        # dryrun = True: "Send" button clicked on confirm page
        # dryrun = False: "Preview" button clicked on first Mailtool page
        dryrun = (submit_button_label != "send")

        recipients = [int(rec) for rec in form.recipients.data.split(",")]
        bodies = []
        participants = []

        try:
            for recipient in recipients:
                prt = lp(recipient)  # type: Participant
                ee = prt.extras

                if ee:
                    e = ee[0]
                    pay_now, pay_to_hotel, items = extras_cost(e)
                else:
                    e, pay_now, pay_to_hotel, items = None, None, None, None

                bodies.append(render_template_string(form.body.data, prt=prt, extras=e,
                                                     pay_now=pay_now, pay_to_hotel=pay_to_hotel, items=items))
                participants.append(prt)

            emails = ehbmail.send(recipients, form.subject.data, bodies, "Mail Tool (%s)" %
                                  current_user.id, replyto=form.replyto.data, dryrun=dryrun, delay=delay_between_messages)

        except Exception as e:
            logger().error("Mailtool: Exception while attempting to send emails: %s" % repr(e))
            flash("An error occurred while rendering or sending your emails: %s. Please check the Mail Archive to see what emails were actually sent." % str(e))
            prts = session.query(Participant).all()
            return render_template("mailtool.html", title="Mail Tool", form=form, participants=prts, delay=delay_between_messages)

        if dryrun:
            return render_template("confirm_emails.html", title="Mail Tool", emails_with_participants=zip(emails, participants), form=form, num_emails=len(emails))

        else:
            return render_template("admin.html", message="%d email(s) were sent." % (len(emails)))

    else:
        prts = session.query(Participant).all()
        return render_template("mailtool.html", title="Mail Tool", form=form, participants=prts, delay=delay_between_messages)


@app.route("/mailarchive.html")
@login_required
def mailarchive():
    all_emails = session.query(Email).order_by(Email.timestamp.desc()).all()
    prts = [lp(em.recipient) for em in all_emails]

    return render_template("email_list.html", title="Mail Archive", message="All sent email messages, in descending order of time they were sent:", emails_with_participants=zip(all_emails, prts))


prt_select = [(str(prt.id), prt.fullname()) for prt in session.query(Participant).all()]
_taf = {"rows": "20", "cols": "80"}


class MailtoolForm(Form):
    recipients = HiddenField()
    subject = StringField("Subject", render_kw={"placeholder": "Enter the email subject"})
    replyto = StringField(
        "Reply-To", render_kw={"placeholder": "Enter an alternative reply-to address (optional)"})
    dryrun = BooleanField("Dry-run")
    body = TextAreaField("Body", render_kw=_taf)


@app.route("/offline-participants.html", methods=["GET", ])
@login_required
def offline_participants():
    form = XlsUploadForm()
    return render_template("offline-participants.html", form=form)


@app.route("/offline-participants.html", methods=["POST", ])
@login_required
def upload_offline_participants():
    form = XlsUploadForm(request.form)

    if form.validate():
        remote_file = request.files[form.file.name]

        if not remote_file.filename.lower().endswith(".xlsx"):
            form.file.errors.append("Please specify an XLSX file.")
            return render_template("offline-participants.html", form=form)

        tmpname = tempfile.mkstemp(suffix=".xlsx")[1]
        remote_file.save(tmpname)

        update_table_from_spreadsheet(Participant, tmpname)

        os.remove(tmpname)

        return render_template("admin.html", message="Database updated.")

    else:
        return render_template("offline-participants.html", form=form)


def send_table_spreadsheet(table, date_fields, money_fields):
    # prepare Excel file
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    bold = workbook.add_format({'bold': True})
    df = workbook.add_format({'num_format': 'dd/mm/yy'})
    # money = workbook.add_format({'num_format': '€ 0'})
    worksheet = workbook.add_worksheet()

    fields = [column.key for column in table.__table__.columns]
    date_fields = set(date_fields)
    money_fields = set(money_fields)

    for i, f in enumerate(fields):
        worksheet.write(0, i, f, bold)

    for i, prt in enumerate(session.query(Participant)):
        row = i + 1

        for col, f in enumerate(fields):
            if f in date_fields:
                worksheet.write(row, col, getattr(prt, f), df)
            else:
                worksheet.write(row, col, getattr(prt, f))

    workbook.close()
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def update_table_from_spreadsheet(table, xlsx_file):
    # assumes that table has an id field, which is in the first column of the Excel table

    wb = load_workbook(xlsx_file)
    sheet = wb.active

    fields = [column.key for column in table.__table__.columns]
    row = 2

    rows_in_db = {row.id: row for row in session.query(table)}

    # make backup of table, just in case
    pickle_filename = "backup_%s_%s.p" % (table.__table__.name, str(datetime.datetime.now()))
    pickle_filename = pickle_filename.replace(" ", "_")
    pickle.dump(rows_in_db, open(pickle_filename, "wb"))

    while True:
        id = sheet.cell(row=row, column=1).value
        if not id:
            # break at first row with empty ID
            break

        obj = rows_in_db[id]
        for i, f in enumerate(fields):
            column = i + 1
            value = sheet.cell(row=row, column=i + 1).value
            setattr(obj, f, value)

        row += 1

    session.commit()


@app.route("/participants.xlsx")
@login_required
def participants_spreadsheet():
    return send_table_spreadsheet(Participant, ["application_time"], ["donation"])


class XlsUploadForm(Form):
    file = FileField("XLSX file")  # , validators=[FileRequired()])


########### RANDOM QUARTETS #############

@app.route("/randomquartets.xlsx")
@login_required
def generate_random_quartets():
    tenors, leads, baris, basses = [session.query(Participant).filter(
        Participant.final_part == p).all() for p in (1, 2, 3, 4)]

    num_quartets = max(len(tenors), len(leads), len(baris), len(basses))
    tenors, leads, baris, basses = [pad(people, num_quartets)
                                    for people in (tenors, leads, baris, basses)]

    songs = [title for (k, title, key, start) in read_songs()]
    songs = pad_songs(songs, num_quartets)

    # generate spreadsheet
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    bold = workbook.add_format({'bold': True})
    italic = workbook.add_format({'italic': True})
    worksheet = workbook.add_worksheet()

    worksheet.set_column(1, 4, 20)
    worksheet.set_column(5, 6, 30)

    worksheet.write(0, 0, "Nr", bold)
    worksheet.write(0, 1, "Tenor", bold)
    worksheet.write(0, 2, "Lead", bold)
    worksheet.write(0, 3, "Bari", bold)
    worksheet.write(0, 4, "Bass", bold)
    worksheet.write(0, 5, "Song", bold)
    worksheet.write(0, 6, "Quartet name", bold)
    worksheet.write(0, 7, "Judge 1", bold)
    worksheet.write(0, 8, "Judge 2", bold)
    worksheet.write(0, 9, "Judge 3", bold)
    worksheet.write(0, 10, "Judge 4", bold)
    worksheet.write(0, 11, "Average", bold)

    for row in range(1, num_quartets + 1):
        worksheet.write(row, 0, row)
        worksheet.write(row, 1, tenors[row - 1].fullname())
        worksheet.write(row, 2, leads[row - 1].fullname())
        worksheet.write(row, 3, baris[row - 1].fullname())
        worksheet.write(row, 4, basses[row - 1].fullname())
        worksheet.write(row, 5, songs[row - 1])
        # formula = '=IFERROR(AVERAGEIF(H%d:K%d;">0"); "---")' % (row+1, row+1)
        # formula = "=2+3"
        formula = '=IFERROR(AVERAGE(H%d:K%d), 0)' % (row + 1, row + 1)
        worksheet.write_formula(row, 11, formula)

    workbook.close()
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def shuffle_until_first_is_not(L, forbidden_first_element):
    random.shuffle(L)

    while L[0] == forbidden_first_element:
        random.shuffle(L)


def pad_songs(songs, num_quartets):
    ret = list()

    # extend with as many copies of the complete song list as will fit
    while len(ret) <= num_quartets - len(songs):
        if ret:
            shuffle_until_first_is_not(songs, ret[-1])
        else:
            random.shuffle(songs)

        ret.extend(songs)

    # extend up to len(songs)-1 many with individual songs from the list
    extras = list(songs)
    if ret:
        shuffle_until_first_is_not(extras, ret[-1])
    else:
        random.shuffle(extras)

    while len(ret) < num_quartets:
        ret.append(extras.pop())

    return ret


def pad(people, size):
    ret = list(people)
    random.shuffle(ret)

    extras = list(people)
    shuffle_until_first_is_not(extras, ret[-1])

    while len(ret) < size:
        ret.append(extras.pop())

    return ret


@app.route("/discount_codes.html")
@login_required
def discountgenerator():
    form = DiscountForm(request.form)
    return render_template("discount_codes.html", title="Generate Discount Code", form=form)


@app.route("/googlemap.html")
@login_required
def mappage():
    return template.render(latlongs=ll, names=nn, parts=pp)
